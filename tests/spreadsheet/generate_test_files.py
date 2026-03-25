"""Generate test spreadsheet files for edge-case testing.

Run:  uv run python tests/spreadsheet/generate_test_files.py
"""

from pathlib import Path
import openpyxl
from openpyxl.cell.cell import TYPE_STRING

OUT = Path(__file__).parent / "fixtures"


def _set_str(ws, row: int, col: int, value: str) -> None:
    """Write a value as an explicit string cell (avoids openpyxl error-type auto-detection)."""
    cell = ws.cell(row, col, value)
    cell.data_type = TYPE_STRING


def _save(wb: openpyxl.Workbook, name: str) -> None:
    path = OUT / name
    wb.save(path)
    print(f"  ✓ {path.name}  ({path.stat().st_size:,} bytes)")


# ─────────────────────────────────────────────────────────────────────────────
#  1. All sheets empty (triggers EmptySpreadsheetError → PDF fallback)
# ─────────────────────────────────────────────────────────────────────────────
def gen_all_empty():
    wb = openpyxl.Workbook()
    wb.active.title = "Sheet1"
    wb.create_sheet("Sheet2")
    wb.create_sheet("Sheet3")
    _save(wb, "all_sheets_empty.xlsx")


# ─────────────────────────────────────────────────────────────────────────────
#  2. Mix: one empty sheet + one with data
# ─────────────────────────────────────────────────────────────────────────────
def gen_mixed_empty():
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Empty"

    ws2 = wb.create_sheet("HasData")
    ws2.append(["Name", "Value"])
    ws2.append(["Alice", "100"])
    ws2.append(["Bob", "200"])
    _save(wb, "mixed_empty_and_data.xlsx")


# ─────────────────────────────────────────────────────────────────────────────
#  3. Massive NaN flood — small file, huge markdown output
#     (tests excel_max_output_ratio guard)
# ─────────────────────────────────────────────────────────────────────────────
def gen_nan_flood():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "NaNFlood"
    cols = 20
    rows = 500
    ws.append([f"Col{i}" for i in range(cols)])
    for _ in range(rows):
        ws.append(["nan"] * cols)
    _save(wb, "nan_flood.xlsx")


# ─────────────────────────────────────────────────────────────────────────────
#  4. All Excel error types in cells and headers
# ─────────────────────────────────────────────────────────────────────────────
def gen_all_errors():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Errors"
    errors = ["#REF!", "#N/A", "#VALUE!", "#NAME?", "#DIV/0!", "#NULL!", "#NUM!", "nan"]
    # Row 1: headers — force string type to avoid openpyxl 'e' (error) data_type
    headers = ["OK_Header"] + errors[:4]
    for col_idx, val in enumerate(headers, 1):
        _set_str(ws, 1, col_idx, val)
    # Data rows
    for i in range(10):
        row = [f"row{i}"] + [errors[j % len(errors)] for j in range(i, i + 4)]
        for col_idx, val in enumerate(row, 1):
            _set_str(ws, i + 2, col_idx, val)
    _save(wb, "all_error_types.xlsx")


# ─────────────────────────────────────────────────────────────────────────────
#  5. Error values as column headers (tests _rename_error_columns)
# ─────────────────────────────────────────────────────────────────────────────
def gen_error_headers():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ErrHeaders"
    # Force string type for error-like headers
    for col_idx, val in enumerate(["#REF!", "#N/A", "#VALUE!", "Normal"], 1):
        _set_str(ws, 1, col_idx, val)
    for i in range(5):
        ws.append([f"a{i}", f"b{i}", f"c{i}", f"d{i}"])
    _save(wb, "error_headers.xlsx")


# ─────────────────────────────────────────────────────────────────────────────
#  6. Duplicate empty-name columns (triggers the _strip_empty iloc bug)
#     Columns B, D, F have no header → all renamed to "" by col_dic
# ─────────────────────────────────────────────────────────────────────────────
def gen_duplicate_empty_cols():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DupEmptyCols"
    # Row 1 = headers: A has a name, B/D/F are blank, C/E have names
    ws.append(["ID", None, "Name", None, "Score", None])
    ws.append(["1", None, "Alice", None, "95", None])
    ws.append(["2", "extra", "Bob", None, "88", None])
    ws.append(["3", None, "Charlie", None, "72", "note"])
    _save(wb, "duplicate_empty_columns.xlsx")


# ─────────────────────────────────────────────────────────────────────────────
#  7. Sheet with only empty rows (data looks present but is all whitespace)
# ─────────────────────────────────────────────────────────────────────────────
def gen_whitespace_only():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Whitespace"
    ws.append(["A", "B", "C"])
    for _ in range(20):
        ws.append(["", "  ", ""])
    _save(wb, "whitespace_only_rows.xlsx")


# ─────────────────────────────────────────────────────────────────────────────
#  8. Single cell — minimal content
# ─────────────────────────────────────────────────────────────────────────────
def gen_single_cell():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Minimal"
    ws["C5"] = "lonely value"
    _save(wb, "single_cell.xlsx")


# ─────────────────────────────────────────────────────────────────────────────
#  9. Multi-sheet: errors on some sheets, clean on others
# ─────────────────────────────────────────────────────────────────────────────
def gen_multi_sheet_mixed_errors():
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Clean"
    ws1.append(["Product", "Price"])
    ws1.append(["Widget", "9.99"])
    ws1.append(["Gadget", "19.99"])

    ws2 = wb.create_sheet("Broken")
    ws2.append(["Label", "Result"])
    # Force string type for error values
    for row_idx, (label, err) in enumerate(
        [("div_err", "#DIV/0!"), ("name_err", "#NAME?"), ("ref_err", "#REF!")], 2
    ):
        ws2.cell(row_idx, 1, label)
        _set_str(ws2, row_idx, 2, err)

    ws3 = wb.create_sheet("AlsoClean")
    ws3.append(["X", "Y"])
    ws3.append(["1", "2"])
    _save(wb, "multi_sheet_mixed_errors.xlsx")


# ─────────────────────────────────────────────────────────────────────────────
# 10. Wide table — many columns, few rows (stress tabulate formatting)
# ─────────────────────────────────────────────────────────────────────────────
def gen_wide_table():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Wide"
    n_cols = 100
    ws.append([f"H{i}" for i in range(n_cols)])
    for r in range(3):
        ws.append([f"r{r}c{c}" for c in range(n_cols)])
    _save(wb, "wide_100_columns.xlsx")


# ─────────────────────────────────────────────────────────────────────────────
# 11. Multiline cell content + special characters
# ─────────────────────────────────────────────────────────────────────────────
def gen_multiline_cells():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Multiline"
    ws.append(["Description", "Notes"])
    ws.append(["Line1\nLine2\nLine3", "Single line"])
    ws.append(["Normal", "Tab\there\nand\nnewlines"])
    ws.append(["Pipe | char", "Backslash \\ test"])
    ws.append(["Accents: é à ü ñ", "CJK: 日本語テスト"])
    _save(wb, "multiline_and_special_chars.xlsx")


# ─────────────────────────────────────────────────────────────────────────────
# 12. Many empty sheets + one tiny sheet (boundary for EmptySpreadsheetError)
# ─────────────────────────────────────────────────────────────────────────────
def gen_many_empty_one_tiny():
    wb = openpyxl.Workbook()
    wb.active.title = "Empty1"
    for i in range(2, 10):
        wb.create_sheet(f"Empty{i}")
    ws = wb.create_sheet("Tiny")
    ws.append(["x"])
    ws.append(["1"])
    _save(wb, "many_empty_one_tiny.xlsx")


# ─────────────────────────────────────────────────────────────────────────────
# 13. Float NaN from merged cells — pandas produces float('nan') for non-anchor
#     cells even with keep_default_na=False
# ─────────────────────────────────────────────────────────────────────────────
def gen_float_nan_merged():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["Name", "Status", "Value"])
    ws.append(["Alice", "OK", "100"])
    ws.append(["Bob", "OK", "200"])
    ws.append(["Carol", "OK", "300"])
    # Merge B2:B4 — only B2 keeps its value; B3, B4 become float NaN in pandas
    ws.merge_cells("B2:B4")
    _save(wb, "float_nan_merged.xlsx")


# ─────────────────────────────────────────────────────────────────────────────
# 14. Multiple independent tables — each sheet has different schema/topic
# ─────────────────────────────────────────────────────────────────────────────
def gen_multi_table_independent():
    wb = openpyxl.Workbook()

    ws1 = wb.active
    ws1.title = "Employees"
    ws1.append(["ID", "Name", "Department", "Salary"])
    ws1.append(["1", "Alice", "Engineering", "95000"])
    ws1.append(["2", "Bob", "Marketing", "72000"])
    ws1.append(["3", "Carol", "Engineering", "98000"])
    ws1.append(["4", "Dave", "Sales", "65000"])

    ws2 = wb.create_sheet("Products")
    ws2.append(["SKU", "Name", "Price", "Stock"])
    ws2.append(["A001", "Widget", "9.99", "150"])
    ws2.append(["A002", "Gadget", "24.50", "80"])
    ws2.append(["A003", "Doohickey", "3.75", "500"])

    ws3 = wb.create_sheet("Orders")
    ws3.append(["OrderID", "CustomerID", "ProductSKU", "Qty", "Date"])
    ws3.append(["ORD-001", "C100", "A001", "5", "2025-01-15"])
    ws3.append(["ORD-002", "C101", "A003", "12", "2025-01-16"])
    ws3.append(["ORD-003", "C100", "A002", "1", "2025-01-17"])

    _save(wb, "multi_table_independent.xlsx")


# ─────────────────────────────────────────────────────────────────────────────
# 15. Multiple sheets with varying sizes — stress test iteration
# ─────────────────────────────────────────────────────────────────────────────
def gen_multi_table_varying_sizes():
    wb = openpyxl.Workbook()

    # Sheet 1: 1 row
    ws1 = wb.active
    ws1.title = "Tiny"
    ws1.append(["Key", "Value"])
    ws1.append(["version", "1.0"])

    # Sheet 2: 50 rows
    ws2 = wb.create_sheet("Medium")
    ws2.append(["Index", "Data", "Category"])
    for i in range(50):
        ws2.append([str(i), f"item_{i}", f"cat_{i % 5}"])

    # Sheet 3: 3 columns, 5 rows
    ws3 = wb.create_sheet("Small")
    ws3.append(["X", "Y", "Z"])
    for i in range(5):
        ws3.append([str(i * 10), str(i * 20), str(i * 30)])

    _save(wb, "multi_table_varying_sizes.xlsx")


# ─────────────────────────────────────────────────────────────────────────────
# 16. Sheets with special characters in names
# ─────────────────────────────────────────────────────────────────────────────
def gen_multi_table_special_names():
    wb = openpyxl.Workbook()

    ws1 = wb.active
    ws1.title = "Q1 2025 (Jan-Mar)"
    ws1.append(["Month", "Revenue"])
    ws1.append(["Jan", "10000"])
    ws1.append(["Feb", "12000"])
    ws1.append(["Mar", "11500"])

    ws2 = wb.create_sheet("Données & Stats")
    ws2.append(["Metric", "Value"])
    ws2.append(["Mean", "42.5"])
    ws2.append(["Median", "40.0"])

    ws3 = wb.create_sheet("Sheet #3 - Notes")
    ws3.append(["Note"])
    ws3.append(["This is a test"])

    _save(wb, "multi_table_special_names.xlsx")


# ─────────────────────────────────────────────────────────────────────────────
# 17. Sheets with overlapping columns — same column names, different data
# ─────────────────────────────────────────────────────────────────────────────
def gen_multi_table_overlapping_columns():
    wb = openpyxl.Workbook()

    ws1 = wb.active
    ws1.title = "2024"
    ws1.append(["Name", "Score", "Status"])
    ws1.append(["Alice", "85", "Pass"])
    ws1.append(["Bob", "42", "Fail"])

    ws2 = wb.create_sheet("2025")
    ws2.append(["Name", "Score", "Status"])
    ws2.append(["Alice", "92", "Pass"])
    ws2.append(["Bob", "78", "Pass"])

    ws3 = wb.create_sheet("Summary")
    ws3.append(["Year", "AvgScore", "PassRate"])
    ws3.append(["2024", "63.5", "50%"])
    ws3.append(["2025", "85.0", "100%"])

    _save(wb, "multi_table_overlapping_columns.xlsx")


# ─────────────────────────────────────────────────────────────────────────────
# 18. Mix of populated and empty sheets interspersed
# ─────────────────────────────────────────────────────────────────────────────
def gen_multi_table_sparse():
    wb = openpyxl.Workbook()

    ws1 = wb.active
    ws1.title = "Data1"
    ws1.append(["A", "B"])
    ws1.append(["1", "2"])

    wb.create_sheet("EmptyA")  # empty

    ws3 = wb.create_sheet("Data2")
    ws3.append(["C", "D"])
    ws3.append(["3", "4"])

    wb.create_sheet("EmptyB")  # empty

    wb.create_sheet("EmptyC")  # empty

    ws6 = wb.create_sheet("Data3")
    ws6.append(["E", "F", "G"])
    ws6.append(["5", "6", "7"])
    ws6.append(["8", "9", "10"])

    _save(wb, "multi_table_sparse.xlsx")


# ─────────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    print("Generating test spreadsheets in:", OUT)
    gen_all_empty()
    gen_mixed_empty()
    gen_nan_flood()
    gen_all_errors()
    gen_error_headers()
    gen_duplicate_empty_cols()
    gen_whitespace_only()
    gen_single_cell()
    gen_multi_sheet_mixed_errors()
    gen_wide_table()
    gen_multiline_cells()
    gen_many_empty_one_tiny()
    gen_float_nan_merged()
    gen_multi_table_independent()
    gen_multi_table_varying_sizes()
    gen_multi_table_special_names()
    gen_multi_table_overlapping_columns()
    gen_multi_table_sparse()
    print("Done.")
